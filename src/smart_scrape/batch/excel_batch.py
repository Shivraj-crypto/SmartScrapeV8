from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from itertools import chain
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

COMMON_URL_HEADERS = (
    "url",
    "urls",
    "link",
    "links",
    "website",
    "web_url",
    "page_url",
    "source_url",
    "deal_url",
    "merchant_url",
)


class BatchInputError(Exception):
    """Raised when the Excel batch input is invalid."""


@dataclass(slots=True)
class ExcelURLRecord:
    row_number: int
    url: str


@dataclass(slots=True)
class BatchSummaryRecord:
    row_number: int
    requested_url: str
    status: str
    final_url: str | None = None
    title: str | None = None
    status_code: int | None = None
    elapsed_ms: int | None = None
    overall_confidence: float | None = None
    candidate_count: int = 0
    used_llm_fallback: bool = False
    fallback_error: str | None = None
    top_offer: str | None = None
    top_offer_type: str | None = None
    html_path: str | None = None
    text_path: str | None = None
    deals_path: str | None = None
    error: str | None = None


def _normalize_header(value: object) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower())
    return normalized.strip("_")


def _looks_like_url(value: str) -> bool:
    candidate = value.strip()
    if not candidate:
        return False

    if "://" not in candidate:
        candidate = f"https://{candidate}"

    parsed = urlparse(candidate)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def load_urls_from_excel(
    input_excel_file: Path,
    url_column: str | None = None,
    sheet_name: str | None = None,
) -> list[ExcelURLRecord]:
    input_path = input_excel_file.expanduser().resolve()
    if not input_path.exists() or not input_path.is_file():
        raise BatchInputError(f"Excel file not found: {input_path}")

    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise BatchInputError(
            "Only .xlsx and .xlsm files are supported for batch scraping."
        )

    workbook = load_workbook(filename=str(input_path), read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise BatchInputError(
                    f"Worksheet '{sheet_name}' was not found in {input_path.name}."
                )
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active

        rows = worksheet.iter_rows(values_only=True)
        first_row = next(rows, None)
        if first_row is None:
            raise BatchInputError("The Excel file is empty.")

        header_map = {
            _normalize_header(value): index
            for index, value in enumerate(first_row)
            if value is not None and _normalize_header(value)
        }

        start_row = 2
        url_index: int | None = None

        if url_column:
            normalized_column = _normalize_header(url_column)
            if normalized_column not in header_map:
                available_columns = ", ".join(sorted(header_map)) or "(none)"
                raise BatchInputError(
                    f"URL column '{url_column}' was not found. "
                    f"Available columns: {available_columns}"
                )
            url_index = header_map[normalized_column]
            row_iterable = rows
        else:
            for candidate in COMMON_URL_HEADERS:
                if candidate in header_map:
                    url_index = header_map[candidate]
                    break

            if url_index is None:
                first_row_urls = [
                    (index, str(value).strip())
                    for index, value in enumerate(first_row)
                    if value is not None and str(value).strip()
                ]
                url_cells = [
                    (index, value)
                    for index, value in first_row_urls
                    if _looks_like_url(value)
                ]
                if not url_cells:
                    raise BatchInputError(
                        "Could not detect a URL column. Use --excel-url-column to set it explicitly."
                    )
                url_index = url_cells[0][0]
                start_row = 1
                row_iterable = chain([first_row], rows)
            else:
                row_iterable = rows

        records: list[ExcelURLRecord] = []
        current_row_number = start_row
        for row_values in row_iterable:
            cell_value = (
                row_values[url_index]
                if url_index < len(row_values) and row_values[url_index] is not None
                else None
            )
            url = str(cell_value).strip() if cell_value is not None else ""
            if url:
                records.append(ExcelURLRecord(row_number=current_row_number, url=url))
            current_row_number += 1

        if not records:
            raise BatchInputError("No URLs were found in the selected worksheet.")

        return records
    finally:
        workbook.close()


def build_output_stem(row_number: int, url: str) -> str:
    candidate = url.strip()
    parsed = urlparse(candidate if "://" in candidate else f"https://{candidate}")
    slug_source = f"{parsed.netloc}{parsed.path}"
    if parsed.query:
        slug_source = f"{slug_source}_{parsed.query}"
    if not slug_source.strip("_/"):
        slug_source = candidate

    slug = re.sub(r"[^a-z0-9]+", "_", slug_source.lower()).strip("_")
    if not slug:
        slug = "url"

    return f"row_{row_number:05d}_{slug[:80]}"


def write_batch_summary_csv(
    records: list[BatchSummaryRecord],
    output_csv_file: Path,
) -> None:
    output_path = output_csv_file.expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "row_number",
        "requested_url",
        "status",
        "final_url",
        "title",
        "status_code",
        "elapsed_ms",
        "overall_confidence",
        "candidate_count",
        "used_llm_fallback",
        "fallback_error",
        "top_offer",
        "top_offer_type",
        "html_path",
        "text_path",
        "deals_path",
        "error",
    ]

    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "row_number": record.row_number,
                    "requested_url": record.requested_url,
                    "status": record.status,
                    "final_url": record.final_url or "",
                    "title": record.title or "",
                    "status_code": record.status_code if record.status_code is not None else "",
                    "elapsed_ms": record.elapsed_ms if record.elapsed_ms is not None else "",
                    "overall_confidence": (
                        f"{record.overall_confidence:.2f}"
                        if record.overall_confidence is not None
                        else ""
                    ),
                    "candidate_count": record.candidate_count,
                    "used_llm_fallback": "yes" if record.used_llm_fallback else "no",
                    "fallback_error": record.fallback_error or "",
                    "top_offer": record.top_offer or "",
                    "top_offer_type": record.top_offer_type or "",
                    "html_path": record.html_path or "",
                    "text_path": record.text_path or "",
                    "deals_path": record.deals_path or "",
                    "error": record.error or "",
                }
            )
